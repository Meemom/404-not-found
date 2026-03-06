import csv
import io
import json
import uuid
from fastapi import APIRouter, UploadFile, File, Form
from google import genai
from google.genai import types
from openpyxl import load_workbook

from db import supabase


def _is_valid_uuid(val: str) -> bool:
    try:
        uuid.UUID(val)
        return True
    except ValueError:
        return False

router = APIRouter(prefix="/upload", tags=["upload"])
client = genai.Client()

PROMPTS = {
    "suppliers": (
        "Extract supplier data from this document. Return a JSON array with objects "
        "containing: name, country, components (what they supply). "
        "Return ONLY valid JSON, no markdown."
    ),
    "sla": (
        "Extract customer SLA data from this document. Return a JSON array with objects "
        "containing: customer (name), annual_revenue, contract_terms, sla_days (int), "
        "penalty_percent (number), payment_history. "
        "Return ONLY valid JSON, no markdown."
    ),
    "bom": (
        "Extract bill of materials data from this document. Return a JSON array with objects "
        "containing: stock_keeping_unit, description, supplier_name, quantity_per_unit (int), "
        "lead_time_days (int), unit_cost (number or null). "
        "Return ONLY valid JSON, no markdown."
    ),
}

MIME_MAP = {
    ".csv": "text/csv",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".pdf": "application/pdf",
}


def _xlsx_to_csv(file_bytes: bytes) -> bytes:
    """Convert an Excel file to CSV bytes."""
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    output = io.StringIO()
    writer = csv.writer(output)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)
    wb.close()
    return output.getvalue().encode("utf-8")


def extract_with_gemini(file_bytes: bytes, filename: str, doc_type: str) -> list:
    """Send file to Gemini and get structured JSON back."""
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in (".xlsx", ".xls"):
        file_bytes = _xlsx_to_csv(file_bytes)
        mime = "text/csv"
        filename = filename.rsplit(".", 1)[0] + ".csv"
    else:
        mime = MIME_MAP.get(ext, "application/octet-stream")

    # Upload file to Gemini
    uploaded = client.files.upload(
        file=io.BytesIO(file_bytes),
        config=types.UploadFileConfig(display_name=filename, mime_type=mime),
    )

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[uploaded, PROMPTS[doc_type]],
    )

    # Parse the JSON from the response
    text = response.text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0].strip()
    return json.loads(text)


@router.post("/suppliers")
async def upload_suppliers(
    file: UploadFile = File(...),
    company_id: str = Form(...),
  ):
    content = await file.read()
    rows = extract_with_gemini(content, file.filename, "suppliers")

    # Store in Supabase if valid company_id
    if _is_valid_uuid(company_id):
        for row in rows:
            supabase.table("suppliers").insert({
                "company_id": company_id,
                "name": row.get("name", ""),
                "country": row.get("country", ""),
            }).execute()

    return {"extracted": rows, "count": len(rows)}


@router.post("/sla")
async def upload_sla(
    file: UploadFile = File(...),
    company_id: str = Form(...),
):
    content = await file.read()
    rows = extract_with_gemini(content, file.filename, "sla")

    if _is_valid_uuid(company_id):
        for row in rows:
            supabase.table("customers").insert({
                "company_id": company_id,
                "name": row.get("customer", ""),
                "sla_days": row.get("sla_days"),
                "penalty_percent": row.get("penalty_percent"),
            }).execute()

    return {"extracted": rows, "count": len(rows)}


@router.post("/bom")
async def upload_bom(
    file: UploadFile = File(...),
    company_id: str = Form(...),
  ):
    content = await file.read()
    rows = extract_with_gemini(content, file.filename, "bom")

    if _is_valid_uuid(company_id):
        for row in rows:
            supabase.table("parts").insert({
                "company_id": company_id,
                "part_number": row.get("part_number", ""),
                "description": row.get("description", ""),
                "supplier_name": row.get("supplier_name", ""),
                "quantity_per_unit": row.get("quantity_per_unit"),
                "lead_time_days": row.get("lead_time_days"),
                "unit_cost": row.get("unit_cost"),
            }).execute()

    return {"extracted": rows, "count": len(rows)}